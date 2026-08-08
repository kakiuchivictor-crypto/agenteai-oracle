"""Testes da rota /tools/organize-spreadsheet (fora do fluxo de RAG)."""

from __future__ import annotations

import io

import pandas as pd
from fastapi.testclient import TestClient


def _messy_xlsx_bytes() -> bytes:
    df = pd.DataFrame(
        {
            "Nome ": ["  Ana ", "Bruno", "Ana", None],
            "Vendas": [100, 200, 100, None],
        }
    )
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return buffer.getvalue()


def test_organize_spreadsheet_returns_cleaned_file_and_summary(api_client: TestClient) -> None:
    response = api_client.post(
        "/tools/organize-spreadsheet",
        files={
            "file": (
                "vendas.xlsx",
                _messy_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert "duplicada" in data["summary"]
    assert data["file_name"] == "vendas_organizado.xlsx"
    assert data["columns"] == ["Nome", "Vendas"]
    assert data["total_rows"] == 2
    assert len(data["preview_rows"]) == 2
    assert data["chart_base64"] is None  # generate_chart=False por padrao

    import base64

    xlsx_bytes = base64.b64decode(data["file_base64"])
    assert xlsx_bytes[:2] == b"PK"


def test_organize_spreadsheet_generates_chart_only_when_requested(api_client: TestClient) -> None:
    response = api_client.post(
        "/tools/organize-spreadsheet",
        data={"generate_chart": "true"},
        files={
            "file": (
                "vendas.xlsx",
                _messy_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["chart_base64"] is not None


def test_organize_spreadsheet_sorts_when_request_text_asks_for_it(api_client: TestClient) -> None:
    response = api_client.post(
        "/tools/organize-spreadsheet",
        data={"request_text": "organize essa planilha por ordem crescente de vendas"},
        files={
            "file": (
                "vendas.xlsx",
                _messy_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert "Ordenada pela coluna 'Vendas'" in data["summary"]
    vendas_index = data["columns"].index("Vendas")
    assert [row[vendas_index] for row in data["preview_rows"]] == ["100", "200"]


def test_organize_spreadsheet_does_not_sort_without_request_text(api_client: TestClient) -> None:
    response = api_client.post(
        "/tools/organize-spreadsheet",
        files={
            "file": (
                "vendas.xlsx",
                _messy_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    assert "Ordenada" not in response.json()["summary"]


def test_organize_spreadsheet_full_request_matches_column_and_formats_currency(
    api_client: TestClient,
) -> None:
    """Regressao ao vivo: reproduz o pedido real do usuario (ordenar por
    'salario', sem acento, contra a coluna 'Salário', com acento — e pedir
    formatacao em reais)."""
    df = pd.DataFrame({"Nome": ["Carla", "Ana", "Bruno"], "Salário": [5000.0, 1800.0, 3000.0]})
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    response = api_client.post(
        "/tools/organize-spreadsheet",
        data={
            "request_text": (
                "Organize todos os funcionarios em ordem crescente de salario. Preserve todas "
                "as colunas e registros, formate os salarios em reais e devolva uma nova "
                "planilha Excel. Nao altere os valores."
            )
        },
        files={
            "file": (
                "funcionarios.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert "Ordenada pela coluna 'Salário'" in data["summary"]
    assert "formatada como moeda" in data["summary"]
    salario_index = data["columns"].index("Salário")
    assert [row[salario_index] for row in data["preview_rows"]] == ["1800", "3000", "5000"]

    import base64

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data["file_base64"])))
    sheet = workbook.active
    assert sheet.cell(row=2, column=salario_index + 1).number_format == '"R$" #,##0.00'
    # valores originais preservados (so a ordem e a formatacao de exibicao mudam)
    assert sheet.cell(row=2, column=salario_index + 1).value == 1800.0


def test_organize_spreadsheet_rejects_unsupported_extension(api_client: TestClient) -> None:
    response = api_client.post(
        "/tools/organize-spreadsheet",
        files={"file": ("relatorio.pdf", b"conteudo qualquer", "application/pdf")},
    )

    assert response.status_code == 400
    assert response.json()["error_code"] == "invalid_file"


def test_organize_spreadsheet_rejects_content_that_does_not_match_extension(
    api_client: TestClient,
) -> None:
    """Bytes que nao sao um xlsx de verdade sao barrados na validacao de
    upload (assinatura de conteudo) antes mesmo de chegar em
    `clean_spreadsheet` — mesma defesa usada em `/documents/upload`."""
    response = api_client.post(
        "/tools/organize-spreadsheet",
        files={
            "file": (
                "quebrado.xlsx",
                b"isto nao e um xlsx valido",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert response.json()["error_code"] == "invalid_file"
