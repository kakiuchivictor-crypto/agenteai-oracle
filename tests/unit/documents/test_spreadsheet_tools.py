"""Testes da limpeza automatica de planilhas (secao de "organizar planilha"
no chat) — operacoes fixas e deterministicas, sem execucao de codigo."""

from __future__ import annotations

import io

import pandas as pd
import pytest

from app.core.exceptions import CorruptedFileError, InvalidFileError
from app.documents.spreadsheet_tools import (
    build_summary_chart,
    clean_spreadsheet,
    dataframe_to_xlsx_bytes,
)


def _xlsx_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return buffer.getvalue()


def _messy_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Nome ": ["  Ana ", "Bruno", "Ana", "", None, "Carla"],
            " Regiao": ["Sul", "Norte", "Sul", None, None, "Norte"],
            "Vendas": [100, 200, 100, None, None, 150],
            "Vazia": [None, None, None, None, None, None],
        }
    )


def test_clean_spreadsheet_removes_empty_rows_columns_and_duplicates() -> None:
    result = clean_spreadsheet(filename="teste.xlsx", content=_xlsx_bytes(_messy_dataframe()))

    assert result.empty_columns_removed == 1
    assert result.empty_rows_removed == 2
    assert result.duplicate_rows_removed == 1
    assert list(result.dataframe.columns) == ["Nome", "Regiao", "Vendas"]
    assert list(result.dataframe["Nome"]) == ["Ana", "Bruno", "Carla"]


def test_clean_spreadsheet_trims_whitespace_from_headers_and_values() -> None:
    result = clean_spreadsheet(filename="teste.xlsx", content=_xlsx_bytes(_messy_dataframe()))

    assert "Nome " not in result.dataframe.columns
    assert " Regiao" not in result.dataframe.columns
    assert all(not str(v).startswith(" ") and not str(v).endswith(" ") for v in result.dataframe["Nome"])


def test_clean_spreadsheet_summary_mentions_each_change() -> None:
    result = clean_spreadsheet(filename="teste.xlsx", content=_xlsx_bytes(_messy_dataframe()))

    summary = result.summary
    assert "vazia" in summary
    assert "duplicada" in summary
    assert "3 linha(s) x 3 coluna(s)" in summary


def test_clean_spreadsheet_handles_already_clean_data() -> None:
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    result = clean_spreadsheet(filename="limpo.xlsx", content=_xlsx_bytes(df))

    assert result.empty_rows_removed == 0
    assert result.empty_columns_removed == 0
    assert result.duplicate_rows_removed == 0
    assert "nenhuma alteracao necessaria" in result.summary


def test_clean_spreadsheet_reads_csv() -> None:
    csv_bytes = "Nome,Idade\nAna,30\nAna,30\nBruno,40\n".encode("utf-8")
    result = clean_spreadsheet(filename="pessoas.csv", content=csv_bytes)

    assert result.duplicate_rows_removed == 1
    assert result.final_rows == 2


def test_clean_spreadsheet_sorts_by_named_column_ascending() -> None:
    df = pd.DataFrame({"Nome": ["Carla", "Ana", "Bruno"], "Vendas": [30, 100, 50]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="ordene por nome em ordem crescente"
    )

    assert result.sort_column == "Nome"
    assert result.sort_ascending is True
    assert list(result.dataframe["Nome"]) == ["Ana", "Bruno", "Carla"]
    assert "Ordenada pela coluna 'Nome' em ordem crescente" in result.summary


def test_clean_spreadsheet_sorts_by_named_column_descending() -> None:
    df = pd.DataFrame({"Nome": ["Carla", "Ana", "Bruno"], "Vendas": [30, 100, 50]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="ordene por vendas decrescente"
    )

    assert result.sort_column == "Vendas"
    assert result.sort_ascending is False
    assert list(result.dataframe["Vendas"]) == [100, 50, 30]


def test_clean_spreadsheet_sort_without_named_column_uses_first_column() -> None:
    df = pd.DataFrame({"Nome": ["Carla", "Ana", "Bruno"], "Vendas": [30, 100, 50]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="organize por ordem crescente"
    )

    assert result.sort_column == "Nome"
    assert result.sort_column_guessed is True
    assert list(result.dataframe["Nome"]) == ["Ana", "Bruno", "Carla"]
    assert "nenhuma coluna foi identificada" in result.summary


def test_clean_spreadsheet_matches_column_name_ignoring_accents() -> None:
    """Regressao: o pedido do usuario real usava 'salario' (sem acento)
    enquanto a coluna da planilha era 'Salário' (com acento) — o casamento
    por substring exato falhava e caia no fallback para a primeira coluna
    errada."""
    df = pd.DataFrame({"Nome": ["Carla", "Ana"], "Salário": [5000, 1800]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="ordene por salario crescente"
    )

    assert result.sort_column == "Salário"
    assert result.sort_column_guessed is False
    assert list(result.dataframe["Salário"]) == [1800, 5000]


def test_clean_spreadsheet_sorts_numerically_when_column_is_formatted_currency_text() -> None:
    """Regressao: uma coluna de salario ja formatada como texto ('R$
    12.500,50') ordenava alfabeticamente (1.800 < 12.500 < 5.000 como
    texto), nao numericamente."""
    df = pd.DataFrame(
        {"Nome": ["A", "B", "C"], "Salário": ["R$ 5.000,00", "R$ 12.500,50", "R$ 1.800,00"]}
    )
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="ordene por salario crescente"
    )

    assert list(result.dataframe["Salário"]) == ["R$ 1.800,00", "R$ 5.000,00", "R$ 12.500,50"]


def test_clean_spreadsheet_applies_currency_format_when_column_is_numeric() -> None:
    df = pd.DataFrame({"Nome": ["Ana"], "Salário": [5000.0]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="formate o salario em reais"
    )

    assert result.currency_format_column == "Salário"
    assert result.warnings == []
    assert "formatada como moeda" in result.summary
    # valores originais preservados, nada foi alterado
    assert list(result.dataframe["Salário"]) == [5000.0]


def test_clean_spreadsheet_warns_when_currency_column_is_not_numeric() -> None:
    df = pd.DataFrame({"Nome": ["Ana"], "Salário": ["R$ 5.000,00"]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="formate o salario em reais"
    )

    assert result.currency_format_column is None
    assert len(result.warnings) == 1
    assert "não estão em formato numérico" in result.warnings[0]
    assert "⚠️" in result.summary


def test_clean_spreadsheet_warns_when_currency_format_column_cannot_be_identified() -> None:
    df = pd.DataFrame({"Nome": ["Ana"], "Valor": [100]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="formate isso em reais por favor"
    )

    assert result.currency_format_column is None
    assert len(result.warnings) == 1
    assert "Não identifiquei qual coluna" in result.warnings[0]


def test_clean_spreadsheet_does_not_sort_without_sort_keywords() -> None:
    df = pd.DataFrame({"Nome": ["Carla", "Ana", "Bruno"], "Vendas": [30, 100, 50]})
    result = clean_spreadsheet(
        filename="t.xlsx", content=_xlsx_bytes(df), request_text="so organize essa planilha"
    )

    assert result.sort_column is None
    assert list(result.dataframe["Nome"]) == ["Carla", "Ana", "Bruno"]
    assert "Ordenada" not in result.summary


def test_clean_spreadsheet_rejects_unsupported_extension() -> None:
    with pytest.raises(InvalidFileError):
        clean_spreadsheet(filename="documento.pdf", content=b"conteudo qualquer")


def test_clean_spreadsheet_rejects_corrupted_file() -> None:
    with pytest.raises(CorruptedFileError):
        clean_spreadsheet(filename="quebrado.xlsx", content=b"isto nao e um xlsx valido")


def test_dataframe_to_xlsx_bytes_produces_a_valid_workbook() -> None:
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    xlsx_bytes = dataframe_to_xlsx_bytes(df)

    assert xlsx_bytes[:2] == b"PK"  # xlsx e um zip
    roundtrip = pd.read_excel(io.BytesIO(xlsx_bytes), engine="openpyxl")
    assert list(roundtrip.columns) == ["A", "B"]
    assert len(roundtrip) == 2


def test_dataframe_to_xlsx_bytes_applies_currency_number_format() -> None:
    import openpyxl

    df = pd.DataFrame({"Nome": ["Ana", "Bruno"], "Salário": [5000.0, 1800.0]})
    xlsx_bytes = dataframe_to_xlsx_bytes(df, currency_columns=["Salário"])

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    # linha 1 e cabecalho; dados comecam na linha 2. Coluna 'Salário' e a 2a (indice 1).
    assert sheet.cell(row=2, column=2).number_format == '"R$" #,##0.00'
    assert sheet.cell(row=3, column=2).number_format == '"R$" #,##0.00'
    # valores numericos preservados exatamente
    assert sheet.cell(row=2, column=2).value == 5000.0
    assert sheet.cell(row=3, column=2).value == 1800.0


def test_build_summary_chart_returns_png_when_numeric_column_exists() -> None:
    df = pd.DataFrame({"Regiao": ["Sul", "Norte", "Sul"], "Vendas": [100, 200, 50]})
    chart_bytes = build_summary_chart(df)

    assert chart_bytes is not None
    assert chart_bytes[:8] == b"\x89PNG\r\n\x1a\n"


def test_build_summary_chart_returns_none_without_numeric_column() -> None:
    df = pd.DataFrame({"Nome": ["Ana", "Bruno"], "Cidade": ["SP", "RJ"]})
    assert build_summary_chart(df) is None


def test_build_summary_chart_returns_none_for_empty_dataframe() -> None:
    df = pd.DataFrame({"Valor": pd.Series(dtype=float)})
    assert build_summary_chart(df) is None
