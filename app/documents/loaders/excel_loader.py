"""Carregador de planilhas Excel (secao 5.3 do prompt mestre).

Cada linha vira uma secao independente e legivel (cabecalho: valor), evitando
transformar a planilha inteira em um unico bloco de texto — exigencia
explicita do prompt mestre.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import CorruptedFileError
from app.documents.loaders.base import DocumentLoader
from app.schemas.document import DocumentSection, ExtractedDocument, ExtractionWarning


def _row_to_text(sheet_name: str, row_number: int, headers: list[str], values: tuple) -> str:
    lines = [f"Planilha: {sheet_name}", f"Linha {row_number}"]
    for header, value in zip(headers, values, strict=False):
        if value is None or str(value).strip() == "":
            continue
        label = header if header else "Coluna sem cabecalho"
        lines.append(f"{label}: {value}")
    return "\n".join(lines)


class ExcelLoader(DocumentLoader):
    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() == ".xlsx"

    def load(self, file_path: Path) -> ExtractedDocument:
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except (InvalidFileException, KeyError, OSError) as exc:
            raise CorruptedFileError(
                f"Nao foi possivel abrir a planilha: {file_path.name}"
            ) from exc

        sections: list[DocumentSection] = []
        warnings: list[ExtractionWarning] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                warnings.append(
                    ExtractionWarning(
                        code="empty_sheet",
                        message="Planilha vazia.",
                        location=f"planilha '{sheet_name}'",
                    )
                )
                continue

            headers = [str(cell) if cell is not None else "" for cell in header_row]
            row_number = 1
            data_rows_found = False

            for values in rows_iter:
                row_number += 1
                if values is None or all(v is None for v in values):
                    continue
                data_rows_found = True
                text = _row_to_text(sheet_name, row_number, headers, values)
                sections.append(
                    DocumentSection(text=text, sheet_name=sheet_name, row_number=row_number)
                )

            if not data_rows_found:
                warnings.append(
                    ExtractionWarning(
                        code="sheet_without_data_rows",
                        message="Planilha possui apenas cabecalho, sem linhas de dados.",
                        location=f"planilha '{sheet_name}'",
                    )
                )

        workbook.close()

        return ExtractedDocument(
            source_path=str(file_path),
            document_format="xlsx",
            sections=sections,
            warnings=warnings,
            raw_metadata={"sheet_count": str(len(workbook.sheetnames))},
        )
